#!/usr/bin/env python3
"""Example: Pull Critical+High vulns from last 14 days, report by repo, export to Excel."""

import sys
import os
from datetime import datetime, timezone
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add parent dir to path so we can import the SDK
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from armorcode import ArmorCodeClient


def epoch_to_str(epoch_ms):
    if not epoch_ms:
        return ""
    try:
        dt = datetime.fromtimestamp(epoch_ms / 1000, tz=timezone.utc)
        return dt.strftime("%Y-%m-%d %H:%M UTC")
    except (ValueError, TypeError, OSError):
        return ""


def write_excel(findings, repos, severities, days_back, output_path):
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    sev_fills = {
        "Critical": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        "High": PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid"),
        "Medium": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
        "Low": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    }
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "ArmorCode Vulnerability Report by Repository"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A3"] = f"Date range: Last {days_back} days | Severities: {', '.join(severities)}"

    headers = ["Repository", "Critical", "High", "Medium", "Low", "Info", "Total"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=col, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin_border
        c.alignment = Alignment(horizontal="center")

    repo_sev = defaultdict(Counter)
    for f in findings:
        repo = (f.get("subProduct") or {}).get("name", "(unmapped)")
        sev_raw = (f.get("severity") or "Unknown").capitalize()
        repo_sev[repo][sev_raw] += 1

    for row, repo in enumerate(repos, 6):
        counts = repo_sev.get(repo, Counter())
        vals = [repo, counts["Critical"], counts["High"], counts["Medium"],
                counts["Low"], counts["Info"], sum(counts.values())]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = thin_border

    ws.column_dimensions["A"].width = 35
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # --- Detail ---
    ws2 = wb.create_sheet("Findings Detail")
    detail_headers = [
        "Repository", "Severity", "Status", "Title", "CVE", "CWE",
        "Source Tool", "Scan Type", "Category", "Component", "Version",
        "Fix Version", "File Path", "Developer", "Found On", "Last Seen",
        "CVSS", "Risk Score", "SLA Breached", "Team", "Product",
        "Environment", "Finding URL",
    ]
    for col, h in enumerate(detail_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font, c.fill, c.border = header_font, header_fill, thin_border

    for row, f in enumerate(findings, 2):
        sev_raw = f.get("severity", "")
        sev = sev_raw.capitalize() if sev_raw else ""
        vals = [
            (f.get("subProduct") or {}).get("name", ""),
            sev,
            f.get("status", ""),
            f.get("title", ""),
            ", ".join(f.get("cve", []) or []),
            f.get("cwe", ""),
            f.get("source", ""),
            ", ".join(f.get("scanType", []) or []),
            f.get("category", ""),
            f.get("componentName", ""),
            f.get("componentVersion", ""),
            f.get("componentFixVersions", ""),
            f.get("filePath", ""),
            f.get("developer", ""),
            epoch_to_str(f.get("foundOn")),
            epoch_to_str(f.get("lastSeenDate")),
            f.get("baseScore", ""),
            f.get("riskScore", ""),
            "Yes" if f.get("slaBreached") else "No",
            (f.get("team") or {}).get("name", ""),
            (f.get("product") or {}).get("name", ""),
            (f.get("environment") or {}).get("name", ""),
            f.get("findingUrl", ""),
        ]
        for col, v in enumerate(vals, 1):
            if isinstance(v, (list, tuple)):
                v = ", ".join(str(x) for x in v)
            c = ws2.cell(row=row, column=col, value=v if v else "")
            c.border = thin_border
            if col == 2 and sev in sev_fills:
                c.fill = sev_fills[sev]
                if sev == "Critical":
                    c.font = Font(color="FFFFFF", bold=True)

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}1"

    wb.save(output_path)
    print(f"Saved: {output_path} ({len(findings)} rows)")


if __name__ == "__main__":
    from collections import defaultdict

    # Connect using env file
    env_path = os.environ.get("AC_ENV", os.path.join(os.path.dirname(__file__), "..", "env"))
    ac = ArmorCodeClient.from_env(env_path)

    # 1. Bulk pull
    severities = ["Critical", "High"]
    days_back = 14
    findings = ac.get_findings(
        severities=severities,
        days_back=days_back,
        dump_path="findings.json",
    )
    print(f"Pulled {len(findings)} findings")

    # 2. List repos
    repo_counts = ac.list_repos()
    print(f"\n{len(repo_counts)} repos:")
    for name, count in repo_counts:
        print(f"  {name}: {count}")

    # 3. Pick top 3 repos (or use CLI args)
    if len(sys.argv) > 1:
        target_repos = sys.argv[1:]
    else:
        target_repos = [name for name, _ in repo_counts[:3]]
    print(f"\nTarget repos: {', '.join(target_repos)}")

    # 4. Get findings per repo and write Excel
    filtered = []
    for repo in target_repos:
        rf = ac.get_findings_by_repo(repo)
        print(f"  {repo}: {len(rf)} findings")
        filtered.extend(rf)

    output = f"vuln_by_repo_{datetime.now().strftime('%Y%m%d')}.xlsx"
    write_excel(filtered, target_repos, severities, days_back, output)
